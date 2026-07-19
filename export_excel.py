#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Export comptable mensuel (.xlsx) — TRABELSI Bijouterie.

Génère un classeur Excel pour un mois donné, avec 7 onglets :
  1. RÉSUMÉ            — synthèse du mois (CA, bénéfice, trésorerie)
  2. VENTES            — détail ligne par ligne
  3. CRÉDITS CLIENTS   — créances (ouvertes en premier)
  4. PAIEMENTS REÇUS   — encaissements du mois (trésorerie réelle)
  5. STOCK             — état au jour de l'export
  6. FOURNISSEURS      — dettes fournisseurs
  7. CHÈQUES           — suivi des encaissements
"""

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

MOIS_FR = ["", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
           "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]

# ── Styles ───────────────────────────────────────────────────────────────────
OR_FONCE = "8A6D1F"
OR_CLAIR = "F5EBD0"
GRIS     = "6B7280"
VERT     = "15803D"
ROUGE    = "B42318"

F_TITRE   = Font(bold=True, size=15, color=OR_FONCE)
F_ENTETE  = Font(bold=True, size=10, color="FFFFFF")
F_GRAS    = Font(bold=True)
F_MUTED   = Font(size=9, color=GRIS)
FILL_ENT  = PatternFill("solid", fgColor=OR_FONCE)
FILL_TOT  = PatternFill("solid", fgColor=OR_CLAIR)
CENTRE    = Alignment(horizontal="center", vertical="center")
_thin     = Side(style="thin", color="D9D9D9")
BORD      = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

FMT_MAD = '#,##0 "MAD"'
FMT_GRS = '#,##0.00 "grs"'
FMT_CTS = '#,##0.00'
FMT_PCT = '0.0"%"'


def _entetes(ws, cols, ligne=1):
    """Écrit la ligne d'en-tête et fige les volets."""
    for i, (titre, largeur, _fmt) in enumerate(cols, start=1):
        c = ws.cell(row=ligne, column=i, value=titre)
        c.font, c.fill, c.alignment, c.border = F_ENTETE, FILL_ENT, CENTRE, BORD
        ws.column_dimensions[get_column_letter(i)].width = largeur
    ws.freeze_panes = ws.cell(row=ligne + 1, column=1)
    ws.auto_filter.ref = f"A{ligne}:{get_column_letter(len(cols))}{ligne}"


def _ligne(ws, r, valeurs, cols, gras=False):
    for i, v in enumerate(valeurs, start=1):
        c = ws.cell(row=r, column=i, value=v)
        c.border = BORD
        fmt = cols[i - 1][2]
        if fmt:
            c.number_format = fmt
        if gras:
            c.font, c.fill = F_GRAS, FILL_TOT
    return r + 1


def _couleur_montant(ws, r, col, valeur):
    if valeur is None:
        return
    ws.cell(row=r, column=col).font = Font(bold=True, color=VERT if valeur >= 0 else ROUGE)


# ── Helpers données ──────────────────────────────────────────────────────────
def _f(v):
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def _du_mois(d, mois):
    return str(d or "")[:7] == mois


def _type_lisible(v):
    return {"service": "Service", "reparation": "Réparation",
            "reprise": "Reprise"}.get(v.get("type_vente") or "produit", "Produit")


def _anciennete(date_str):
    try:
        d = datetime.strptime(str(date_str)[:10], "%Y-%m-%d")
        return (datetime.now() - d).days
    except Exception:
        return None


# ── Onglets ──────────────────────────────────────────────────────────────────
def _onglet_resume(wb, mois, ventes_mois, ventes_prec, credits, articles, is_lot):
    ws = wb.create_sheet("RÉSUMÉ")
    if mois == "tout":
        ws["A1"] = "TRABELSI Joaillerie — HISTORIQUE COMPLET"
    else:
        an, m = mois.split("-")
        ws["A1"] = f"TRABELSI Joaillerie — {MOIS_FR[int(m)]} {an}"
    ws["A1"].font = F_TITRE
    ws["A2"] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    ws["A2"].font = F_MUTED
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18

    def bloc(titre, r):
        c = ws.cell(row=r, column=1, value=titre)
        c.font, c.fill = F_ENTETE, FILL_ENT
        ws.cell(row=r, column=2).fill = FILL_ENT
        ws.cell(row=r, column=3).fill = FILL_ENT
        return r + 1

    def kv(r, label, val, fmt=FMT_MAD, couleur=False):
        ws.cell(row=r, column=1, value=label).border = BORD
        c = ws.cell(row=r, column=2, value=val)
        c.number_format, c.border = fmt, BORD
        if couleur and isinstance(val, (int, float)):
            c.font = Font(bold=True, color=VERT if val >= 0 else ROUGE)
        return r + 1

    par_type = {}
    for v in ventes_mois:
        t = _type_lisible(v)
        d = par_type.setdefault(t, {"ca": 0.0, "benef": 0.0, "nb": 0})
        d["ca"] += _f(v.get("pv"))
        d["benef"] += _f(v.get("benef"))
        d["nb"] += 1

    ca_tot = sum(d["ca"] for d in par_type.values())
    bn_tot = sum(d["benef"] for d in par_type.values())
    or_vendu = sum(_f(v.get("or_grs")) for v in ventes_mois)

    r = 4
    r = bloc("CHIFFRE D'AFFAIRES", r)
    for t in ("Produit", "Service", "Réparation", "Reprise"):
        if t in par_type:
            r = kv(r, f"  {t}", round(par_type[t]["ca"]))
    r = kv(r, "TOTAL CA", round(ca_tot))
    ws.cell(row=r - 1, column=1).font = F_GRAS
    ws.cell(row=r - 1, column=2).font = F_GRAS
    ws.cell(row=r - 1, column=1).fill = FILL_TOT
    ws.cell(row=r - 1, column=2).fill = FILL_TOT

    r += 1
    r = bloc("BÉNÉFICE", r)
    for t in ("Produit", "Service", "Réparation", "Reprise"):
        if t in par_type:
            r = kv(r, f"  {t}", round(par_type[t]["benef"]), couleur=True)
    r = kv(r, "BÉNÉFICE NET", round(bn_tot), couleur=True)
    ws.cell(row=r - 1, column=1).font = F_GRAS
    ws.cell(row=r - 1, column=1).fill = FILL_TOT
    ws.cell(row=r - 1, column=2).fill = FILL_TOT
    marge = (bn_tot / ca_tot * 100) if ca_tot else 0
    r = kv(r, "Marge nette", round(marge, 1), FMT_PCT)

    r += 1
    r = bloc("ACTIVITÉ", r)
    r = kv(r, "Nombre de ventes", len(ventes_mois), "0")
    r = kv(r, "Clients différents", len({(v.get("client") or "?").strip().lower()
                                         for v in ventes_mois if v.get("client")}), "0")
    r = kv(r, "OR vendu", round(or_vendu, 2), FMT_GRS)

    # Trésorerie : encaissements du mois (paiements sur crédits)
    encaisse = 0.0
    for c in credits:
        for p in c.get("paiements", []) or []:
            if mois == "tout" or _du_mois(p.get("date"), mois):
                encaisse += _f(p.get("montant"))
    reste_total = sum(_f(c.get("reste")) for c in credits
                      if (c.get("statut") or "") in ("rien", "avance"))
    r += 1
    r = bloc("TRÉSORERIE & CRÉANCES", r)
    r = kv(r, "Encaissé (sur crédits)" if mois == "tout" else "Encaissé ce mois (sur crédits)", round(encaisse))
    r = kv(r, "Total dû par les clients", round(reste_total))
    r = kv(r, "Nombre de créances ouvertes",
           len([c for c in credits if (c.get("statut") or "") in ("rien", "avance")]), "0")

    # Comparaison mois précédent
    ca_prec = sum(_f(v.get("pv")) for v in ventes_prec)
    bn_prec = sum(_f(v.get("benef")) for v in ventes_prec)
    r += 1
    if mois != "tout":
        r = bloc("COMPARAISON MOIS PRÉCÉDENT", r)
        r = kv(r, "CA mois précédent", round(ca_prec))
        r = kv(r, "Évolution CA", round((ca_tot - ca_prec) / ca_prec * 100, 1) if ca_prec else 0,
               FMT_PCT, couleur=True)
        r = kv(r, "Bénéfice mois précédent", round(bn_prec))
        r = kv(r, "Évolution bénéfice", round((bn_tot - bn_prec) / bn_prec * 100, 1) if bn_prec else 0,
               FMT_PCT, couleur=True)

    # Stock (au jour de l'export)
    q = lambda a: int(a.get("quantite") or 1)
    mlt = lambda a: 1 if is_lot(a) else q(a)
    r += 1
    r = bloc(f"STOCK AU {datetime.now().strftime('%d/%m/%Y')}", r)
    r = kv(r, "Valeur du stock (prix de revient)",
           round(sum(_f(a.get("pa")) * mlt(a) for a in articles)))
    r = kv(r, "OR en stock", round(sum(_f(a.get("or_grs")) * mlt(a) for a in articles), 2), FMT_GRS)
    r = kv(r, "Articles en stock (hors chaînes)",
           sum(q(a) for a in articles if not is_lot(a)), "0")
    r = kv(r, "Chaînes en stock", sum(q(a) for a in articles if is_lot(a)), "0")
    return ws


def _bloc_par_mois(ws, r, ventes):
    """Tableau récapitulatif mois par mois (export historique)."""
    entetes = ["Mois", "Ventes", "CA", "Bénéfice", "Marge %", "OR vendu (grs)"]
    largeurs = [16, 10, 16, 16, 11, 15]
    for i, (t, w) in enumerate(zip(entetes, largeurs), start=1):
        c = ws.cell(row=r, column=i, value=t)
        c.font, c.fill, c.alignment, c.border = F_ENTETE, FILL_ENT, CENTRE, BORD
        if ws.column_dimensions[get_column_letter(i)].width < w:
            ws.column_dimensions[get_column_letter(i)].width = w
    r += 1
    par_mois = {}
    for v in ventes:
        m = str(v.get("date_vente") or "")[:7]
        if not m:
            continue
        d = par_mois.setdefault(m, {"nb": 0, "ca": 0.0, "bn": 0.0, "or": 0.0})
        d["nb"] += 1
        d["ca"] += _f(v.get("pv"))
        d["bn"] += _f(v.get("benef"))
        d["or"] += _f(v.get("or_grs"))
    t_nb = t_ca = t_bn = t_or = 0
    for m in sorted(par_mois):
        d = par_mois[m]
        t_nb += d["nb"]; t_ca += d["ca"]; t_bn += d["bn"]; t_or += d["or"]
        an, mm = m.split("-")
        vals = [f"{MOIS_FR[int(mm)]} {an}", d["nb"], round(d["ca"]), round(d["bn"]),
                round(d["bn"] / d["ca"] * 100, 1) if d["ca"] else None, round(d["or"], 2)]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = BORD
            c.number_format = [None, "0", FMT_MAD, FMT_MAD, FMT_PCT, FMT_CTS][i - 1] or "General"
        _couleur_montant(ws, r, 4, d["bn"])
        r += 1
    vals = ["TOTAL", t_nb, round(t_ca), round(t_bn),
            round(t_bn / t_ca * 100, 1) if t_ca else None, round(t_or, 2)]
    for i, v in enumerate(vals, start=1):
        c = ws.cell(row=r, column=i, value=v)
        c.border, c.font, c.fill = BORD, F_GRAS, FILL_TOT
        c.number_format = [None, "0", FMT_MAD, FMT_MAD, FMT_PCT, FMT_CTS][i - 1] or "General"
    return r + 1


def _onglet_ventes(wb, ventes_mois):
    cols = [("Date", 12, None), ("Réf", 10, None), ("Article", 24, None), ("Type", 12, None),
            ("OR (grs)", 11, FMT_CTS), ("Diam. (cts)", 11, FMT_CTS), ("Émer. (cts)", 11, FMT_CTS),
            ("Rubis (cts)", 11, FMT_CTS), ("Saphir (cts)", 11, FMT_CTS),
            ("Prix revient", 14, FMT_MAD), ("Prix vente", 14, FMT_MAD),
            ("Bénéfice", 14, FMT_MAD), ("Marge %", 10, FMT_PCT),
            ("Client", 24, None), ("Paiement", 14, None)]
    ws = wb.create_sheet("VENTES")
    _entetes(ws, cols)
    r = 2
    t_pa = t_pv = t_bn = t_or = 0.0
    for v in sorted(ventes_mois, key=lambda x: str(x.get("date_vente") or "")):
        pa, pv, bn = _f(v.get("pa")), _f(v.get("pv")), _f(v.get("benef"))
        t_pa += pa; t_pv += pv; t_bn += bn; t_or += _f(v.get("or_grs"))
        d = str(v.get("date_vente") or "")[:10]
        try:
            d = datetime.strptime(d, "%Y-%m-%d").strftime("%d/%m/%Y")
        except Exception:
            pass
        r = _ligne(ws, r, [
            d, (v.get("ref") or ""), v.get("article") or "", _type_lisible(v),
            _f(v.get("or_grs")) or None, _f(v.get("d")) or None, _f(v.get("em")) or None,
            _f(v.get("r")) or None, _f(v.get("s")) or None,
            pa or None, pv or None, bn,
            round(bn / pv * 100, 1) if pv else None,
            v.get("client") or "", v.get("mode_paiement") or "",
        ], cols)
        _couleur_montant(ws, r - 1, 12, bn)
    _ligne(ws, r, ["TOTAUX", "", f"{len(ventes_mois)} vente(s)", "", round(t_or, 2),
                   None, None, None, None, round(t_pa), round(t_pv), round(t_bn),
                   round(t_bn / t_pv * 100, 1) if t_pv else None, "", ""], cols, gras=True)
    return ws


def _onglet_credits(wb, credits):
    cols = [("Client", 26, None), ("Contact", 16, None), ("Date achat", 12, None),
            ("Réfs", 18, None), ("Article", 22, None), ("Montant total", 14, FMT_MAD),
            ("Déjà payé", 14, FMT_MAD), ("Reste dû", 14, FMT_MAD),
            ("Statut", 12, None), ("Ancienneté (j)", 13, "0")]
    ws = wb.create_sheet("CRÉDITS CLIENTS")
    _entetes(ws, cols)
    ouverts = [c for c in credits if (c.get("statut") or "") in ("rien", "avance")]
    soldes = [c for c in credits if c not in ouverts]
    ouverts.sort(key=lambda c: str(c.get("date_achat") or ""))
    r, t_reste = 2, 0.0
    for c in ouverts + soldes:
        paye = sum(_f(p.get("montant")) for p in c.get("paiements", []) or [])
        reste = _f(c.get("reste"))
        if (c.get("statut") or "") in ("rien", "avance"):
            t_reste += reste
        r = _ligne(ws, r, [
            c.get("client") or "", c.get("contact") or "", c.get("date_achat") or "",
            str(c.get("refs") or ""), c.get("article") or "",
            _f(c.get("montant_total")), paye, reste,
            {"rien": "Impayé", "avance": "Avance", "solde": "Soldé"}.get(c.get("statut"), c.get("statut") or ""),
            _anciennete(c.get("date_achat")),
        ], cols)
    _ligne(ws, r, ["TOTAL CRÉANCES OUVERTES", "", "", "", "", None, None, round(t_reste), "", None],
           cols, gras=True)
    return ws


def _onglet_paiements(wb, credits, mois):
    cols = [("Date", 12, None), ("Client", 26, None), ("Montant", 14, FMT_MAD),
            ("Mode", 14, None), ("Crédit n°", 10, "0"), ("Article", 24, None)]
    ws = wb.create_sheet("PAIEMENTS REÇUS")
    _entetes(ws, cols)
    lignes = []
    for c in credits:
        for p in c.get("paiements", []) or []:
            if mois is None or _du_mois(p.get("date"), mois):
                lignes.append((p.get("date"), c.get("client") or "", _f(p.get("montant")),
                               p.get("mode") or "", c.get("id"), c.get("article") or ""))
    lignes.sort(key=lambda x: str(x[0]))
    r, tot = 2, 0.0
    for l in lignes:
        tot += l[2]
        r = _ligne(ws, r, list(l), cols)
    _ligne(ws, r, ["TOTAL ENCAISSÉ", f"{len(lignes)} paiement(s)", round(tot), "", None, ""],
           cols, gras=True)
    return ws


def _onglet_stock(wb, articles, is_lot):
    cols = [("Réf", 14, None), ("Article", 24, None), ("Quantité", 10, "0"),
            ("OR (grs)", 11, FMT_CTS), ("Prix revient", 14, FMT_MAD),
            ("Diam.", 9, FMT_CTS), ("Émer.", 9, FMT_CTS), ("Rubis", 9, FMT_CTS),
            ("Saphir", 9, FMT_CTS), ("Date entrée", 12, None)]
    ws = wb.create_sheet("STOCK")
    ws.cell(row=1, column=1, value=f"Stock au {datetime.now().strftime('%d/%m/%Y')}").font = F_TITRE
    _entetes(ws, cols, ligne=3)
    r, t_or, t_val, t_q = 4, 0.0, 0.0, 0
    for a in sorted(articles, key=lambda x: str(x.get("article") or "")):
        q = int(a.get("quantite") or 1)
        mlt = 1 if is_lot(a) else q
        t_or += _f(a.get("or_grs")) * mlt
        t_val += _f(a.get("pa")) * mlt
        t_q += q
        r = _ligne(ws, r, [
            str(a.get("ref_code") or a.get("id")), a.get("article") or "", q,
            _f(a.get("or_grs")) or None, _f(a.get("pa")) or None,
            _f(a.get("d")) or None, _f(a.get("em")) or None,
            _f(a.get("r")) or None, _f(a.get("s")) or None,
            str(a.get("date") or "")[:10],
        ], cols)
    _ligne(ws, r, ["TOTAUX", f"{len(articles)} référence(s)", t_q, round(t_or, 2),
                   round(t_val), None, None, None, None, ""], cols, gras=True)
    return ws


def _onglet_fournisseurs(wb, fournisseurs):
    cols = [("Fournisseur", 24, None), ("Contact", 16, None), ("Date commande", 13, None),
            ("N° commande", 14, None), ("Article", 22, None), ("Montant", 14, FMT_MAD),
            ("Payé", 14, FMT_MAD), ("Reste dû", 14, FMT_MAD), ("Statut", 12, None)]
    ws = wb.create_sheet("FOURNISSEURS")
    _entetes(ws, cols)
    r, tot = 2, 0.0
    for f in fournisseurs:
        paye = sum(_f(p.get("montant")) for p in f.get("paiements", []) or [])
        reste = _f(f.get("reste"))
        if (f.get("statut") or "") in ("rien", "avance"):
            tot += reste
        r = _ligne(ws, r, [
            f.get("fournisseur") or "", f.get("contact") or "", f.get("date_commande") or "",
            f.get("num_commande") or "", f.get("article") or "",
            _f(f.get("montant_total")), paye, reste,
            {"rien": "Impayé", "avance": "Avance", "solde": "Soldé"}.get(f.get("statut"), f.get("statut") or ""),
        ], cols)
    _ligne(ws, r, ["TOTAL DÛ FOURNISSEURS", "", "", "", "", None, None, round(tot), ""],
           cols, gras=True)
    return ws


def _onglet_cheques(wb, cheques):
    cols = [("Client", 24, None), ("Montant", 14, FMT_MAD), ("Banque", 16, None),
            ("N° chèque", 16, None), ("Date chèque", 12, None),
            ("Date encaissement", 15, None), ("Statut", 14, None), ("Réf article", 14, None)]
    ws = wb.create_sheet("CHÈQUES")
    _entetes(ws, cols)
    r, tot = 2, 0.0
    for c in cheques:
        tot += _f(c.get("montant"))
        r = _ligne(ws, r, [
            c.get("client") or "", _f(c.get("montant")), c.get("banque") or "",
            str(c.get("numero") or ""), c.get("date_cheque") or "",
            c.get("date_encaissement") or "", c.get("statut") or "",
            str(c.get("ref_article") or ""),
        ], cols)
    _ligne(ws, r, ["TOTAL", round(tot), "", "", "", "", "", ""], cols, gras=True)
    return ws


# ── Point d'entrée ───────────────────────────────────────────────────────────
def mois_precedent(mois):
    an, m = int(mois[:4]), int(mois[5:7])
    return f"{an - 1}-12" if m == 1 else f"{an}-{m - 1:02d}"


def nom_fichier(mois, ventes=None):
    if mois == "tout":
        m = sorted({str(v.get("date_vente"))[:7] for v in (ventes or [])
                    if v.get("date_vente")})
        debut = m[0] if m else "debut"
        return f"TRABELSI_HISTORIQUE_{debut}_a_{datetime.now().strftime('%Y-%m')}.xlsx"
    an, mm = mois.split("-")
    return f"TRABELSI_{mois}_{MOIS_FR[int(mm)]}.xlsx"


def generer(mois, ventes, credits, articles, fournisseurs, cheques, is_lot):
    """Construit le classeur ('AAAA-MM' ou 'tout') et retourne les octets .xlsx."""
    tout = (mois == "tout")
    if tout:
        ventes_sel = [v for v in ventes if v.get("date_vente")]
        ventes_prec = []
    else:
        ventes_sel = [v for v in ventes if _du_mois(v.get("date_vente"), mois)]
        ventes_prec = [v for v in ventes if _du_mois(v.get("date_vente"), mois_precedent(mois))]

    wb = Workbook()
    wb.remove(wb.active)
    ws = _onglet_resume(wb, mois, ventes_sel, ventes_prec, credits, articles, is_lot)
    if tout:
        # Récapitulatif mois par mois, à la suite du résumé
        r = ws.max_row + 3
        ws.cell(row=r, column=1, value="DÉTAIL MOIS PAR MOIS").font = F_TITRE
        _bloc_par_mois(ws, r + 1, ventes_sel)
    _onglet_ventes(wb, ventes_sel)
    _onglet_credits(wb, credits)
    _onglet_paiements(wb, credits, None if tout else mois)
    _onglet_stock(wb, articles, is_lot)
    _onglet_fournisseurs(wb, fournisseurs)
    _onglet_cheques(wb, cheques)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
